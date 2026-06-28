"""PII sanitization helpers for uploaded reporting files.

The upload pipeline must sanitize data before it is queryable or sent to an
LLM-backed step. If a project-specific script is configured, this module runs
that first. A deterministic fallback sanitizer is kept here so uploads are not
accepted without a privacy pass.
"""

from __future__ import annotations

import csv
import hashlib
import re
import subprocess
import sys
import tempfile
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


PII_COLUMN_KEYWORDS = {
    "address",
    "adresse",
    "birth",
    "cin",
    "commcare",
    "contact",
    "dob",
    "email",
    "id_patient",
    "identifier",
    "link",
    "name",
    "nom",
    "patient_id",
    "phone",
    "photo",
    "prenom",
    "prénom",
    "surname",
    "telephone",
    "téléphone",
    "url",
}

PSEUDONYMIZE_COLUMN_KEYWORDS = {"record_id", "source_row_id"}

PII_VALUE_PATTERNS = [
    re.compile(r"\b[\w.+-]+@[\w-]+(?:\.[\w-]+)+\b"),
    re.compile(r"(?:\+?\d[\s().-]*){8,}"),
    re.compile(r"https?://\S+|www\.\S+", re.IGNORECASE),
]


@dataclass
class SanitizationReport:
    original_filename: str
    sanitized_filename: str
    row_count: int
    original_columns: list[str]
    retained_columns: list[str]
    removed_columns: list[str]
    pseudonymized_columns: list[str] = field(default_factory=list)
    redacted_cells: int = 0
    external_script_used: bool = False


def sanitize_uploaded_csv(
    input_path: Path,
    output_path: Path,
    *,
    original_filename: str,
    script_path: str | None = None,
) -> SanitizationReport:
    """Sanitize a CSV upload and write the sanitized copy to ``output_path``."""
    if script_path:
        script = Path(script_path).expanduser()
        if script.exists():
            with tempfile.TemporaryDirectory() as tmp:
                script_output = Path(tmp) / output_path.name
                _run_external_script(script, input_path, script_output)
                report = _sanitize_csv_builtin(
                    script_output,
                    output_path,
                    original_filename=original_filename,
                )
                report.external_script_used = True
                return report

    return _sanitize_csv_builtin(
        input_path,
        output_path,
        original_filename=original_filename,
    )


def convert_xlsx_to_csv(input_path: Path, output_path: Path) -> str:
    """Convert the first non-empty XLSX worksheet to CSV.

    Returns the worksheet name used for conversion.
    """
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        worksheet = _first_non_empty_sheet(workbook)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with output_path.open("w", encoding="utf-8", newline="") as target:
            writer = csv.writer(target)
            for row in worksheet.iter_rows(values_only=True):
                writer.writerow([_cell_to_csv_value(cell) for cell in row])
        return worksheet.title
    finally:
        workbook.close()


def _run_external_script(script: Path, input_path: Path, output_path: Path) -> None:
    """Run a configured PII script.

    Expected script interface:
      python script.py <input_csv> <output_csv>
    """
    subprocess.run(
        [sys.executable, str(script), str(input_path), str(output_path)],
        check=True,
        capture_output=True,
        text=True,
    )


def _first_non_empty_sheet(workbook):
    for worksheet in workbook.worksheets:
        if worksheet.max_row > 1 or worksheet.max_column > 1 or worksheet["A1"].value is not None:
            return worksheet
    return workbook.worksheets[0]


def _cell_to_csv_value(value) -> str:
    if value is None:
        return ""
    return str(value)


def _sanitize_csv_builtin(
    input_path: Path,
    output_path: Path,
    *,
    original_filename: str,
) -> SanitizationReport:
    with input_path.open("r", encoding="utf-8-sig", newline="") as source:
        reader = csv.DictReader(source)
        original_columns = reader.fieldnames or []
        removed_columns = [c for c in original_columns if _is_pii_column(c)]
        pseudonymized_columns = [
            c
            for c in original_columns
            if c not in removed_columns and _is_pseudonymized_column(c)
        ]
        retained_columns = [c for c in original_columns if c not in removed_columns]

        output_path.parent.mkdir(parents=True, exist_ok=True)
        row_count = 0
        redacted_cells = 0

        with output_path.open("w", encoding="utf-8", newline="") as target:
            writer = csv.DictWriter(target, fieldnames=retained_columns)
            writer.writeheader()

            for row in reader:
                cleaned: dict[str, str] = {}
                for column in retained_columns:
                    value = row.get(column, "")
                    if column in pseudonymized_columns:
                        cleaned[column] = _stable_token(column, value)
                    else:
                        redacted, changed = _redact_pii_values(value)
                        cleaned[column] = redacted
                        if changed:
                            redacted_cells += 1
                writer.writerow(cleaned)
                row_count += 1

    return SanitizationReport(
        original_filename=original_filename,
        sanitized_filename=output_path.name,
        row_count=row_count,
        original_columns=original_columns,
        retained_columns=retained_columns,
        removed_columns=removed_columns,
        pseudonymized_columns=pseudonymized_columns,
        redacted_cells=redacted_cells,
    )


def _is_pii_column(column: str) -> bool:
    normalized = _normalize_column(column)
    return any(keyword in normalized for keyword in PII_COLUMN_KEYWORDS)


def _is_pseudonymized_column(column: str) -> bool:
    normalized = _normalize_column(column)
    return any(keyword in normalized for keyword in PSEUDONYMIZE_COLUMN_KEYWORDS)


def _normalize_column(column: str) -> str:
    return re.sub(r"[^a-z0-9_]+", "_", column.lower())


def _redact_pii_values(value: str) -> tuple[str, bool]:
    cleaned = value
    for pattern in PII_VALUE_PATTERNS:
        cleaned = pattern.sub("[REDACTED]", cleaned)
    return cleaned, cleaned != value


def _stable_token(column: str, value: str) -> str:
    if not value:
        return ""
    digest = hashlib.sha256(f"{column}:{value}".encode("utf-8")).hexdigest()[:16]
    return f"anon_{digest}"
